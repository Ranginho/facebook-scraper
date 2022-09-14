import json
import random
import numpy as np
import time
import openpyxl
from typing import Optional

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions


def start_browser(driver_address, headless=True):
    chrome_options = Options()
    chrome_options = webdriver.ChromeOptions()
    prefs = {"profile.default_content_setting_values.notifications" : 2}
    chrome_options.add_experimental_option("prefs", prefs)
    
    if headless:
        chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(driver_address, options=chrome_options)
    return driver


def wait_to_load(browser, el_type, el_value):
    try:
        WebDriverWait(browser, 10).until(expected_conditions.presence_of_element_located((el_type, el_value)))
        return True
    except TimeoutException:
        return False
    
def save_info(info, info_dir, filename):
    info_dir.mkdir(parents=True, exist_ok=True)
    
    filename = info_dir / f"{filename}.json"
    
    with open(filename, 'w') as f:
        f.write(json.dumps(info, default=str))
        
        
def save_post(post, post_dir):
    save_info(post, post_dir, post['post_id'])
     
    
def save_image(image_info, image_dir):
    save_info(image_info['images_lowquality'], image_dir/'images', image_info['post_id'])
    save_info(image_info['images_lowquality_description'], image_dir/'descriptions', image_info['post_id'])


def save_profile_info(profile_info, profile_dir):
    save_info(profile_info, profile_dir, "profile")

        
# generate random numbers for sleep after every request    
# TODO: change low_out, high_out
def generate_sleep_time_dist(low=500, high=1000, size=1000, low_out=5000, hight_out=6000, size_out_pct=0.01):
    size_out = int(size * size_out_pct)

    sleep_times = np.round(np.random.uniform(low=low, high=high, size=size))
    sleep_times_out = np.round(np.random.uniform(low=low_out, high=hight_out, size=size_out))

    return np.concatenate((sleep_times, sleep_times_out))


def log_into_facebook(browser, email, password):    
    browser.get("https://www.facebook.com/")
    
    # politeness
    time.sleep(random.randint(2, 5))

    wait_to_load(browser, By.ID, "email")
    browser.find_element(By.ID, "email").send_keys(email)

    wait_to_load(browser, By.ID, "pass")
    browser.find_element(By.ID, "pass").send_keys(password)

    wait_to_load(browser, By.NAME, "login")
    browser.find_element(By.NAME, "login").click()
    
    # politeness
    time.sleep(random.randint(1, 3))
    

def like_random_thing(browser, url):    
    browser.get(url)
    
    # politeness
    time.sleep(random.randint(2, 5))
    
    # browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    
    body_size = browser.find_element(By.TAG_NAME, "body").size
    browser.set_window_size(body_size["width"], body_size["height"])
    time.sleep(1)
    
    if wait_to_load(browser, By.CSS_SELECTOR, "[aria-label=Like]"):
        likes = browser.find_elements(By.CSS_SELECTOR, "[aria-label=Like]")
        random.choice(likes).click()
        
        
def extract_id(friend_dict: dict) -> str:
    """
    Args:
        friend_dict(Dict) - Dictionary of a single friend of user

    Return:
        string - id of that friend
    """

    return friend_dict['id']


def read_accounts(ACCOUNTS_PATH):
    wb_obj = openpyxl.load_workbook(ACCOUNTS_PATH)
    sheet_obj = wb_obj.active
    num_rows = sheet_obj.max_row
    profile_ids = []
    for i in range(num_rows):
        cell_obj = sheet_obj.cell(row = i+1, column = 1)
        profile_ids.append(cell_obj.value)
    
    return profile_ids


class CallBackException(Exception):
    """exception to interrupt post scraping"""
    pass


class SearchPagePersistor:
    search_page_url: Optional[str] = None
    handle_lock = False
    
    def get_current_search_page(self) -> Optional[str]:
        return self.search_page_url

    def set_search_page(self, page_url: str) -> None:
        self.search_page_url = page_url
        
        if self.handle_lock:
            self.handle_lock = False
            raise CallBackException # to change account
        else:
            self.handle_lock = True
            